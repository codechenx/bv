import abc
import csv
import gzip
import re

from .config import TypeConfig
from .data import Data
from .utils import detect_file_encoding, tabstr_to_tab


class Reader:
    """
    file reader
    :ivar _path: str, file path
    :ivar _data: Data(), data class sotore file data
    :ivar _config: file type configure
    """

    def __init__(self):
        self._path = None
        self._data = Data()
        self._config = None

    @property
    def path(self):
        return self._path

    @property
    def data(self):
        return self._data

    @property
    def config(self):
        return self._config

    @abc.abstractmethod
    def load_file(self, m_path, m_config=None, compressed=False,
                  m_header=True):
        """
        :param m_path: str, file path
        :param m_config: TypeConfig(), file processing configure
        :param compressed: bool, check if compressed
        :return:
        """
        raise Exception("implement load_file is required")

    def update_config(self, m_config):
        self.load_file(self._path, m_config)


class GeneralReader(Reader):
    def load_file(self, m_path, m_config=None, compressed=False,
                  m_header=True):
        """
        tabular-like file  reader

        :param m_path: str, file path
        :param m_config: TypeConfig(), file processing configure
        :param compressed: bool, check if compressed
        :return:
        """
        assert isinstance(m_path, str) and isinstance(m_config, TypeConfig)

        self._path, self._config = m_path, m_config
        try:
            if compressed:
                encoding = detect_file_encoding(self._path)
                f = gzip.open(self._path)
            else:
                f = open(self._path)
            ignore_regex = self._config.ignore_config.get("regex", None)
            metadata_regex = self._config.metadata_config.get("regex", None)
            sep = self._config.body_config.get("sep", None)
            sep = tabstr_to_tab(sep)  # translate "\\t" to "\t" if possible
            for line in f:
                line = line.decode(
                    encoding).strip() if compressed else line.strip()
                # ignore line
                if ignore_regex and re.match(ignore_regex, line):
                    continue
                # add header data
                if metadata_regex and re.match(metadata_regex, line):
                    self._data.metadata.append(line)
                    continue
                # add header data if True
                if m_header:
                    self.data.header = list(csv.reader([line],
                                                       delimiter=sep))[0]
                    m_header = False
                    continue
                # add body data
                self.data.body.append(
                    list(csv.reader([line], delimiter=sep))[0])
            f.close()
            self.data.covert_possible_col_numberic()
        except FileNotFoundError:
            raise Exception("File dose not exists")


class ExcelReader(Reader):
    def load_file(self, m_path, m_config=None, compressed=False,
                  m_header=True):
        """
        excel file reader

        :param m_path: str, file path
        :param m_config: TypeConfig(), file processing configure
        :param compressed: bool, check if compressed
        :return:
        """
        import openpyxl

        assert isinstance(m_path, str) and isinstance(m_config, TypeConfig)

        self._path, self._config = m_path, m_config
        
        try:
            wb = openpyxl.load_workbook(self._path)
            sheet = wb.get_active_sheet()
            max_col = sheet.max_column
            max_row = sheet.max_row

            for i in range(1, max_row + 1):
                # add header data if True
                if m_header:
                    self.data.header = []
                    for j in range(1, max_col + 1):
                        self.data.header.append(sheet.cell(row=i, column=j).value)
                    m_header = False
                    continue
                # add body data
                row = []
                for j in range(1, max_col + 1):
                    row.append(sheet.cell(row=i, column=j).value)
                self.data.body.append(row)
        except FileNotFoundError:
            raise Exception("File dose not exists")
